import os
import csv
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class SNRCalculatorGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("SNR 信噪比计算工具")
        self.root.geometry("900x650")
        self.root.resizable(True, True)

        self.data_dir = tk.StringVar(value=r"E:\xwechat_files\wxid_tj50m3v5ah4x22_17b3\msg\file\2026-04")
        self.center_freq = tk.IntVar(value=120000)
        self.half_bw = tk.IntVar(value=50)
        self.auto_detect_each = tk.BooleanVar(value=False)
        self.results = []

        self._setup_ui()

    def _setup_ui(self):
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=60)
        title_frame.pack(fill="x")
        title_frame.pack_propagate(False)

        tk.Label(
            title_frame, text="SNR 信噪比计算工具",
            font=("Microsoft YaHei", 18, "bold"),
            bg="#2c3e50", fg="white"
        ).pack(pady=15)

        config_frame = tk.LabelFrame(
            self.root, text=" 参数配置 ",
            font=("Microsoft YaHei", 11),
            padx=15, pady=10
        )
        config_frame.pack(fill="x", padx=15, pady=(15, 5))

        tk.Label(config_frame, text="数据文件夹:", font=("Microsoft YaHei", 10)).grid(row=0, column=0, sticky="w")
        tk.Entry(config_frame, textvariable=self.data_dir, width=55, font=("Microsoft YaHei", 10)).grid(row=0, column=1, padx=5)
        tk.Button(config_frame, text="浏览...", command=self.browse_folder, width=8).grid(row=0, column=2, padx=5)

        freq_frame = tk.Frame(config_frame)
        freq_frame.grid(row=1, column=0, columnspan=3, sticky="w", pady=(10, 0))

        tk.Label(freq_frame, text="信号中心频率 (kHz):", font=("Microsoft YaHei", 10)).pack(side="left")
        tk.Entry(freq_frame, textvariable=self.center_freq, width=12, font=("Microsoft YaHei", 10)).pack(side="left", padx=(5, 10))
        tk.Label(freq_frame, text="带宽一半 (kHz):", font=("Microsoft YaHei", 10)).pack(side="left")
        tk.Entry(freq_frame, textvariable=self.half_bw, width=12, font=("Microsoft YaHei", 10)).pack(side="left", padx=(5, 10))
        tk.Button(freq_frame, text="🔍 自动检测", command=self.do_auto_detect, font=("Microsoft YaHei", 9), width=10, cursor="hand2").pack(side="left", padx=5)
        tk.Checkbutton(freq_frame, text="对每个文件单独检测", variable=self.auto_detect_each, font=("Microsoft YaHei", 9)).pack(side="left", padx=(10, 0))
        self.signal_range_label = tk.Label(freq_frame, text=f"信号区域: [{self.center_freq.get() - self.half_bw.get()}, {self.center_freq.get() + self.half_bw.get()}] kHz", font=("Microsoft YaHei", 9), fg="#7f8c8d")
        self.signal_range_label.pack(side="left", padx=10)

        self.detected_label = tk.Label(self.root, text="", font=("Microsoft YaHei", 9), fg="#27ae60", padx=15, anchor="w")
        self.detected_label.pack(fill="x")

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)

        self.run_btn = tk.Button(
            btn_frame, text="▶ 开始计算", command=self.run_calculation,
            font=("Microsoft YaHei", 11, "bold"),
            bg="#27ae60", fg="white", width=15, height=2,
            cursor="hand2", relief="flat"
        )
        self.run_btn.pack(side="left", padx=10)

        tk.Button(
            btn_frame, text="↻ 重置", command=self.reset,
            font=("Microsoft YaHei", 10),
            width=10, height=2,
            cursor="hand2", relief="flat"
        ).pack(side="left")

        self.save_btn = tk.Button(
            btn_frame, text="💾 保存Excel", command=self.save_to_excel,
            font=("Microsoft YaHei", 10),
            width=12, height=2,
            cursor="hand2", relief="flat", state="disabled"
        )
        self.save_btn.pack(side="left", padx=10)

        tree_frame = tk.Frame(self.root)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(5, 5))

        columns = ("文件名", "信号峰值(dBm)", "噪底均值(dBm)", "SNR(dB)", "检测中心(kHz)")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)

        col_widths = [200, 120, 120, 100, 130]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor="center")

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.tag_configure("even", background="#f8f9fa")
        self.tree.tag_configure("odd", background="white")

        self.status_label = tk.Label(
            self.root, text="就绪", font=("Microsoft YaHei", 10),
            anchor="w", padx=15, pady=5
        )
        self.status_label.pack(fill="x")

        self.progress = ttk.Progressbar(self.root, mode="indeterminate")
        self.progress.pack(fill="x", padx=15, pady=(0, 5))

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.data_dir.set(folder)
            self.update_signal_range_label()

    def update_signal_range_label(self):
        center = self.center_freq.get()
        half_bw = self.half_bw.get()
        self.signal_range_label.config(text=f"信号区域: [{center - half_bw}, {center + half_bw}] kHz")

    def do_auto_detect(self):
        data_dir = self.data_dir.get()
        if not os.path.isdir(data_dir):
            messagebox.showwarning("警告", "请先选择有效的数据文件夹！")
            return

        csv_files = sorted([f for f in os.listdir(data_dir) if f.lower().endswith('.csv')])
        if not csv_files:
            messagebox.showwarning("警告", "未找到 CSV 文件！")
            return

        fpath = os.path.join(data_dir, csv_files[0])
        freqs, powers = self.parse_csv(fpath)
        if freqs is None or len(freqs) == 0:
            messagebox.showwarning("警告", "无法解析 CSV 文件！")
            return

        result = self.auto_detect_signal(freqs, powers)
        if result[0] is None:
            messagebox.showwarning("警告", "未能检测到信号点！")
            return

        center, half_bw = result
        self.center_freq.set(center)
        self.half_bw.set(half_bw)
        self.update_signal_range_label()
        self.detected_label.config(text=f"✅ 已自动检测 | 中心频率: {center} kHz | 带宽: {half_bw*2} kHz | ({csv_files[0]} 等{len(csv_files)}个文件)")

    def reset(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.results = []
        self.status_label.config(text="已重置")
        self.progress.stop()
        self.save_btn.config(state="disabled")
        self.detected_label.config(text="")

    def parse_csv(self, filepath):
        freqs = []
        powers = []
        in_data = False
        try:
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row in reader:
                    if not row:
                        continue
                    if row[0].strip() == 'DATA':
                        in_data = True
                        continue
                    if in_data:
                        try:
                            freq = float(row[0])
                            power = float(row[1])
                            freqs.append(freq)
                            powers.append(power)
                        except (ValueError, IndexError):
                            continue
        except Exception:
            return None, None
        return (np.array(freqs), np.array(powers)) if freqs else (None, None)

    # ------------------- 核心修改部分 -------------------
    def calc_noise_floor(self, noise_powers):
        """只使用平均底噪的算法"""
        return np.mean(noise_powers)
    # ----------------------------------------------------

    def calc_snr(self, freqs, powers, center, half_bw):
        signal_mask = (freqs >= center - half_bw) & (freqs <= center + half_bw)
        noise_mask = ~signal_mask

        signal_powers = powers[signal_mask]
        noise_powers = powers[noise_mask]

        if len(signal_powers) == 0:
            raise ValueError("信号区域内无数据点")
        if len(noise_powers) == 0:
            raise ValueError("噪声区域内无数据点")

        signal_peak_dbm = np.max(signal_powers)
        noise_mean_dbm = self.calc_noise_floor(noise_powers)
        snr_db = signal_peak_dbm - noise_mean_dbm

        return signal_peak_dbm, noise_mean_dbm, snr_db

    def auto_detect_signal(self, freqs, powers):
        noise_floor = self.calc_noise_floor(powers)
        threshold = noise_floor + 6

        above_threshold = powers > threshold
        if not np.any(above_threshold):
            return None, None

        diff = np.diff(above_threshold.astype(int))
        rising = np.where(diff == 1)[0]
        falling = np.where(diff == -1)[0]

        if len(rising) == 0 or len(falling) == 0:
            peak_idx = np.argmax(powers)
            return freqs[peak_idx], 100

        if falling[0] < rising[0]:
            rising = rising[1:]
        if len(rising) == 0 or len(falling) == 0:
            peak_idx = np.argmax(powers)
            return freqs[peak_idx], 100

        peaks = []
        for r, f in zip(rising, falling):
            if f > r:
                segment_powers = powers[r:f+1]
                seg_peak_idx = np.argmax(segment_powers)
                peaks.append((r + seg_peak_idx, np.max(segment_powers)))

        if not peaks:
            peak_idx = np.argmax(powers)
            return freqs[peak_idx], 100

        main_peak_idx = max(peaks, key=lambda x: x[1])[0]
        center_freq = freqs[main_peak_idx]

        signal_mask = powers > threshold
        signal_freqs = freqs[signal_mask]
        half_bw = int((np.max(signal_freqs) - np.min(signal_freqs)) / 2)

        return int(center_freq), max(half_bw, 50)

    def run_calculation(self):
        data_dir = self.data_dir.get()
        center = self.center_freq.get()
        half_bw = self.half_bw.get()
        auto_detect_each = self.auto_detect_each.get()
        self.update_signal_range_label()

        if not os.path.isdir(data_dir):
            messagebox.showerror("错误", f"文件夹不存在:\n{data_dir}")
            return

        csv_files = sorted([f for f in os.listdir(data_dir) if f.lower().endswith('.csv')])
        if not csv_files:
            messagebox.showwarning("警告", "未找到 CSV 文件！")
            return

        self.run_btn.config(state="disabled", text="计算中...")
        self.progress.start(10)
        self.status_label.config(text=f"正在处理 0/{len(csv_files)} 个文件...")
        self.reset()

        def worker():
            results = []
            for i, fname in enumerate(csv_files):
                fpath = os.path.join(data_dir, fname)
                try:
                    freqs, powers = self.parse_csv(fpath)
                    if freqs is None or len(freqs) == 0:
                        self.root.after(0, self.add_result, fname, "解析失败")
                        continue

                    file_center = center
                    file_half_bw = half_bw
                    detect_info = ""
                    if auto_detect_each:
                        detect_result = self.auto_detect_signal(freqs, powers)
                        if detect_result[0] is not None:
                            file_center, file_half_bw = detect_result
                            detect_info = f" (检测: {file_center} kHz)"

                    sig, noise, snr = self.calc_snr(freqs, powers, file_center, file_half_bw)
                    results.append((fname, sig, noise, snr, file_center, file_half_bw))
                    self.root.after(0, self.add_result, fname, None, sig, noise, snr, file_center if auto_detect_each else None)
                    self.root.after(0, self.status_label.config, {"text": f"正在处理 {i+1}/{len(csv_files)} 个文件...{detect_info}"})
                except Exception as e:
                    self.root.after(0, self.add_result, fname, str(e))
                    self.root.after(0, self.status_label.config, {"text": f"处理 {fname} 时出错: {e}"})

            self.root.after(0, self.on_complete, results)

        threading.Thread(target=worker, daemon=True).start()

    def add_result(self, fname, error_msg=None, sig=None, noise=None, snr=None, center=None):
        if error_msg:
            self.tree.insert("", "end", values=(fname, "错误", error_msg, "", ""))
        else:
            center_str = str(center) if center else ""
            self.tree.insert("", "end", values=(fname, f"{sig:.2f}", f"{noise:.2f}", f"{snr:.2f}", center_str))

    def on_complete(self, results):
        self.progress.stop()
        self.run_btn.config(state="normal", text="▶ 开始计算")

        if not results:
            self.status_label.config(text="处理完成，但无有效结果")
            return

        snrs = [r[3] for r in results]
        self.results = results

        self.tree.insert("", "end", values=("-" * 20, "-" * 14, "-" * 14, "-" * 10, "-" * 16))
        self.tree.insert("", "end", values=(
            f"平均值 (共{len(results)}个)",
            f"{np.mean([r[1] for r in results]):.2f}",
            f"{np.mean([r[2] for r in results]):.2f}",
            f"{np.mean(snrs):.2f}",
            ""
        ))
        self.tree.insert("", "end", values=(
            "统计",
            f"最小:{np.min([r[1] for r in results]):.2f}",
            f"最大:{np.max([r[2] for r in results]):.2f}",
            f"SNR范围:{min(snrs):.2f}~{max(snrs):.2f}",
            ""
        ))

        self.status_label.config(text=f"✅ 处理完成！共 {len(results)} 个文件 | SNR 平均: {np.mean(snrs):.2f} dB | 范围: {min(snrs):.2f} ~ {max(snrs):.2f} dB")
        self.save_btn.config(state="normal")

    def save_to_excel(self):
        if not self.results:
            messagebox.showwarning("警告", "没有可保存的数据！")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=f"SNR结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        if not filepath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "SNR计算结果"

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            data_font = Font(size=10)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            headers = ["文件名", "信号峰值(dBm)", "噪底均值(dBm)", "SNR(dB)", "检测中心(kHz)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            for row_idx, result in enumerate(self.results, 2):
                fname, sig, noise, snr = result[0], result[1], result[2], result[3]
                center = result[4] if len(result) > 4 else ""
                values = [fname, round(sig, 2), round(noise, 2), round(snr, 2), center]
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

            snrs = [r[3] for r in self.results]
            summary_row = len(self.results) + 2
            ws.cell(row=summary_row, column=1, value="平均值").font = Font(bold=True, size=10)
            ws.cell(row=summary_row, column=2, value=round(np.mean([r[1] for r in self.results]), 2))
            ws.cell(row=summary_row, column=3, value=round(np.mean([r[2] for r in self.results]), 2))
            ws.cell(row=summary_row, column=4, value=round(np.mean(snrs), 2))

            for col in range(1, 6):
                ws.column_dimensions[chr(64 + col)].width = 20

            wb.save(filepath)
            messagebox.showinfo("成功", f"结果已保存到:\n{filepath}")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败:\n{e}")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = SNRCalculatorGUI()
    app.run()